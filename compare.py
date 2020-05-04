from pathlib import Path
from openpyxl import load_workbook

mappe1 = Path("compare_file1")
mappe2 = Path("compare_file2")
fil1 = list(mappe1.glob("*"))[0]
fil2 = list(mappe2.glob("*"))[0]

bok = load_workbook("mal.xlsx", data_only=True)
bok1 = load_workbook(fil1)
bok2 = load_workbook(fil2)

# Sammenligner først arkene i boken

ark = bok["Arknavn"]
ark["A4"] = fil1.name
ark["B4"] = fil2.name

rad = 5
for arknavn in bok1.sheetnames:
    ark[f"A{rad}"] = arknavn
    rad += 1

rad = 5
for arknavn in bok2.sheetnames:
    ark[f"B{rad}"] = arknavn
    rad += 1

# Sjekker cellegrid 100 * 100 på hvert ark

like_arknavn = [x for x in bok1.sheetnames if x in bok2.sheetnames]

ark = bok["Celler"]
rad = 4
for arknavn in like_arknavn:
    ark1 = bok1[arknavn]
    ark2 = bok2[arknavn]
    for r in range(1, 101):
        for k in range(1, 101):
            celle1 = ark1.cell(r, k).value
            celle2 = ark2.cell(r, k).value
            if celle1 != celle2:
                ark[f"A{rad}"] = arknavn
                ark[f"B{rad}"] = ark1.cell(r, k).coordinate
                ark[f"C{rad}"] = celle1
                ark[f"D{rad}"] = celle2
                rad += 1

bok.save("Rapport.xlsx")
